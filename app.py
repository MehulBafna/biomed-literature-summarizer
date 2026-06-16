"""Gradio app: abstractive summarization of biomedical research papers."""

import tempfile

import gradio as gr
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from document_utils import (
    clean_text,
    extract_text_from_pdf,
    fetch_papers_for_export,
    fetch_pubmed_abstract,
    get_filepath,
)
from summarizer import MODEL_NAME, summarize_document, summarize_paper_for_export, summarize_structured

LENGTH_PRESETS = {
    "Short": {"target_words": 80},
    "Medium": {"target_words": 150},
    "Long": {"target_words": 300},
}

MIN_WORDS = 20

SAMPLE_ABSTRACT = (
    "Background: Sepsis remains a leading cause of mortality in intensive care units "
    "worldwide. Early identification of patients at risk of clinical deterioration is "
    "critical for timely intervention. Methods: We trained a gradient-boosted ensemble "
    "model on vital sign and laboratory data from 12,000 ICU admissions to predict "
    "sepsis onset 6 hours before clinical recognition. Results: The model achieved an "
    "AUROC of 0.89, outperforming the existing SOFA-based screening protocol "
    "(AUROC 0.76). Sensitivity at a fixed 90% specificity threshold was 0.81. "
    "Conclusions: Machine-learning early-warning systems can meaningfully improve "
    "sepsis detection lead time and may reduce time spent manually reviewing charts "
    "during systematic case review."
)

_EXCEL_COLUMNS = [
    ("Title", 55),
    ("Authors", 28),
    ("Year", 7),
    ("Source", 14),
    ("Key Takeaway", 45),
    ("Abstract Summary", 55),
    ("Materials & Methods", 45),
    ("Results", 45),
    ("URL", 38),
]

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_WRAP = Alignment(wrap_text=True, vertical="top")


def estimate_reading_time(text, wpm=200):
    words = len(text.split())
    return words, words / wpm


def format_stats(original_text, summary_text):
    orig_words, orig_minutes = estimate_reading_time(original_text)
    sum_words, sum_minutes = estimate_reading_time(summary_text)
    time_saved_pct = 0.0
    if orig_minutes > 0:
        time_saved_pct = max(0.0, (1 - sum_minutes / orig_minutes) * 100)
    return (
        f"**Original:** {orig_words:,} words (~{orig_minutes:.1f} min read)  \n"
        f"**Summary:** {sum_words:,} words (~{sum_minutes:.1f} min read)  \n"
        f"**Estimated reading-time saved:** {time_saved_pct:.0f}%"
    )


def run_summarization(raw_text, length_choice, structured, progress=None):
    text = clean_text(raw_text)
    if len(text.split()) < MIN_WORDS:
        return "", f"Please provide at least {MIN_WORDS} words to summarize."
    try:
        if structured:
            sections = summarize_structured(text, progress=progress)
            summary = "\n\n".join(f"{label}:\n{body}" for label, body in sections.items())
        else:
            preset = LENGTH_PRESETS[length_choice]
            summary = summarize_document(text, progress=progress, **preset)
    except RuntimeError as exc:
        return "", str(exc)
    return summary, format_stats(text, summary)


def handle_pdf(file, length_choice, structured, progress=gr.Progress()):
    if file is None:
        return "", "", "Please upload a PDF file."
    progress(0, desc="Extracting text from PDF...")
    raw_text = extract_text_from_pdf(get_filepath(file))
    text = clean_text(raw_text)
    if len(text.split()) < MIN_WORDS:
        return "", "", "Could not extract enough text from this PDF."
    preview = text[:1500] + ("…" if len(text) > 1500 else "")
    summary, stats = run_summarization(raw_text, length_choice, structured, progress)
    return preview, summary, stats


def handle_text(text, length_choice, structured, progress=gr.Progress()):
    return run_summarization(text, length_choice, structured, progress)


def handle_pubmed(pmid_or_url, length_choice, structured, progress=gr.Progress()):
    if not pmid_or_url.strip():
        return "", "", "Please enter a PubMed ID or URL."
    try:
        progress(0, desc="Fetching from PubMed...")
        abstract = fetch_pubmed_abstract(pmid_or_url)
    except Exception as exc:
        return "", "", f"Could not fetch PubMed record: {exc}"
    summary, stats = run_summarization(abstract, length_choice, structured, progress)
    return abstract, summary, stats


def _build_excel(papers: list[dict]) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Literature Review"

    for col, (header, width) in enumerate(_EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _WRAP
        ws.column_dimensions[cell.column_letter].width = width

    field_order = [
        "title", "authors", "year", "source",
        "key_takeaway", "abstract_summary", "materials_methods", "results", "url",
    ]
    for row, paper in enumerate(papers, 2):
        for col, field in enumerate(field_order, 1):
            cell = ws.cell(row=row, column=col, value=paper.get(field, ""))
            cell.alignment = _WRAP
        ws.row_dimensions[row].height = 80

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


def handle_keyword_export(keywords, progress=gr.Progress()):
    if not keywords.strip():
        return None, "Please enter keywords to search."
    try:
        progress(0.05, desc="Searching PubMed and Semantic Scholar...")
        papers = fetch_papers_for_export(keywords.strip(), total=15)
        if not papers:
            return None, "No papers found. Try broader or different keywords."

        n = len(papers)
        enriched = []
        for i, paper in enumerate(papers):
            progress(
                0.1 + (i / n) * 0.82,
                desc=f"Summarizing paper {i + 1} of {n}: {paper['title'][:60]}...",
            )
            enriched.append(summarize_paper_for_export(paper))

        progress(0.95, desc="Building Excel file...")
        filepath = _build_excel(enriched)
        progress(1.0, desc="Done!")
        return filepath, f"✓ Found and summarized **{n} papers**. Click below to download."
    except Exception as exc:
        return None, f"Error: {exc}"


with gr.Blocks(title="Biomedical Literature Summarizer") as demo:
    gr.Markdown(
        """
        # 🧬 Biomedical Literature Summarizer

        Abstractive summarization of biomedical research papers using
        **Claude**. Upload a paper, paste text, pull an abstract from PubMed,
        or search by keyword to export a full literature review to Excel.
        """
    )

    length_choice = gr.Radio(
        choices=list(LENGTH_PRESETS.keys()),
        value="Medium",
        label="Summary length",
    )
    structured_toggle = gr.Checkbox(
        label="Structured summary (Abstract / Materials & Methods / Results & Discussion)",
        value=False,
        info="Splits the summary into sections instead of one paragraph "
        "(ignores the length setting above). Works best on full-paper PDFs "
        "and structured abstracts; falls back to a short overview otherwise.",
    )

    with gr.Tabs():
        with gr.Tab("📄 Upload PDF"):
            pdf_file = gr.File(label="Research paper (PDF)", file_types=[".pdf"])
            pdf_btn = gr.Button("Summarize PDF", variant="primary")
            with gr.Accordion("Extracted text preview", open=False):
                pdf_preview = gr.Textbox(show_label=False, lines=8, interactive=False)
            pdf_summary = gr.Textbox(label="Key findings summary", lines=6)
            pdf_stats = gr.Markdown()
            pdf_btn.click(
                handle_pdf,
                [pdf_file, length_choice, structured_toggle],
                [pdf_preview, pdf_summary, pdf_stats],
            )

        with gr.Tab("✏️ Paste Text"):
            text_input = gr.Textbox(
                label="Abstract or full text",
                lines=12,
                placeholder="Paste research paper text here...",
            )
            text_btn = gr.Button("Summarize Text", variant="primary")
            text_summary = gr.Textbox(label="Key findings summary", lines=6)
            text_stats = gr.Markdown()
            gr.Examples(examples=[[SAMPLE_ABSTRACT]], inputs=[text_input])
            text_btn.click(
                handle_text,
                [text_input, length_choice, structured_toggle],
                [text_summary, text_stats],
            )

        with gr.Tab("🔎 PubMed Fetch"):
            pmid_input = gr.Textbox(
                label="PubMed ID or URL",
                placeholder="e.g. 31978945 or https://pubmed.ncbi.nlm.nih.gov/31978945/",
            )
            pmid_btn = gr.Button("Fetch & Summarize", variant="primary")
            pmid_abstract = gr.Textbox(label="Fetched title + abstract", lines=8, interactive=False)
            pmid_summary = gr.Textbox(label="Key findings summary", lines=6)
            pmid_stats = gr.Markdown()
            gr.Examples(examples=[["31978945"]], inputs=[pmid_input])
            pmid_btn.click(
                handle_pubmed,
                [pmid_input, length_choice, structured_toggle],
                [pmid_abstract, pmid_summary, pmid_stats],
            )

        with gr.Tab("🔬 Keyword → Excel"):
            gr.Markdown(
                "Search PubMed and Semantic Scholar for up to 15 relevant papers, "
                "summarize each with Claude, and download the results as an Excel file."
            )
            keyword_input = gr.Textbox(
                label="Research keywords",
                placeholder="e.g.  CRISPR liver cancer gene editing  |  deep learning chest X-ray  |  COVID-19 cytokine storm biomarkers",
                lines=2,
            )
            keyword_btn = gr.Button("Search & Export to Excel", variant="primary")
            keyword_status = gr.Markdown()
            keyword_file = gr.File(label="Download Excel", file_types=[".xlsx"])
            keyword_btn.click(
                handle_keyword_export,
                [keyword_input],
                [keyword_file, keyword_status],
            )

    gr.Markdown(
        f"""
        ---
        **Model:** [`{MODEL_NAME}`](https://www.anthropic.com/claude) via the
        Anthropic API.
        """
    )

if __name__ == "__main__":
    demo.queue().launch()
